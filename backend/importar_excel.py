"""Migra el historial de cuadres desde el Excel original (CUADRE CAJA-1.xlsm)
a PostgreSQL, para que la app continúe la numeración de turnos.

Columnas de la hoja BASE DE DATOS:
    TURNO, FECHA, HORA, USUARIO, JORNADA, VENTAS NETAS, ENTREGAS, TARJETAS,
    NEQUI O QR, VENTAS FE.CREDITO, BASE DEL DIA, BASE ANTERIOR, TOTAL

Uso:
    .venv/bin/python importar_excel.py "../CUADRE CAJA-1.xlsm"
"""
import asyncio
import re
import sys
from datetime import date, datetime, time, timedelta

from openpyxl import load_workbook
from sqlalchemy import select

from database import SessionLocal, init_models
from models import CuadreCaja

JORNADAS_VALIDAS = {"MAÑANA", "TARDE", "TODO EL DIA"}


def parsear_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, (int, float)):
        # Serial de Excel (días desde 1899-12-30)
        return (datetime(1899, 12, 30) + timedelta(days=float(valor))).date()
    return None


def parsear_hora(valor):
    if isinstance(valor, time):
        return valor
    if isinstance(valor, datetime):
        return valor.time()
    if isinstance(valor, (int, float)):
        # Fracción de día de Excel
        segundos = round(float(valor) % 1 * 86400)
        return time(segundos // 3600, (segundos % 3600) // 60, segundos % 60)
    if isinstance(valor, str):
        # "8:45 p. m." / "1:15 p.m" / "11:37 a. m."
        match = re.match(r"\s*(\d{1,2}):(\d{2})\s*([ap])", valor.lower())
        if match:
            hora, minuto, meridiano = int(match.group(1)), int(match.group(2)), match.group(3)
            if meridiano == "p" and hora != 12:
                hora += 12
            if meridiano == "a" and hora == 12:
                hora = 0
            return time(hora % 24, minuto)
    return None


def parsear_entero(valor):
    if valor is None or valor == "":
        return 0
    try:
        return int(round(float(valor)))
    except (ValueError, TypeError):
        return 0


def normalizar_jornada(valor):
    texto = str(valor or "").strip().upper()
    texto = texto.replace("MANANA", "MAÑANA")
    return texto if texto in JORNADAS_VALIDAS else "TODO EL DIA"


async def importar(ruta_excel: str):
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    ws = wb["BASE DE DATOS"]

    filas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        turno = parsear_entero(fila[0])
        fecha = parsear_fecha(fila[1])
        if turno == 0 or fecha is None:
            continue
        filas.append(
            CuadreCaja(
                numero_turno=turno,
                fecha=fecha,
                hora=parsear_hora(fila[2]) or time(0, 0),
                usuario_id=None,
                usuario_nombre=str(fila[3] or "").strip().upper() or None,
                jornada=normalizar_jornada(fila[4]),
                ventas_netas=parsear_entero(fila[5]),
                entregas=parsear_entero(fila[6]),
                tarjetas=parsear_entero(fila[7]),
                bonos=0,
                nequi_qr=parsear_entero(fila[8]),
                fact_electronica_credito=parsear_entero(fila[9]),
                base_del_dia=parsear_entero(fila[10]),
                base_anterior=parsear_entero(fila[11]),
                sobrante_faltante=parsear_entero(fila[12]),
            )
        )

    await init_models()
    async with SessionLocal() as db:
        existentes = await db.execute(select(CuadreCaja.numero_turno))
        turnos_existentes = {t for (t,) in existentes}

        nuevos = [f for f in filas if f.numero_turno not in turnos_existentes]
        db.add_all(nuevos)
        await db.commit()

    omitidos = len(filas) - len(nuevos)
    print(f"Importados {len(nuevos)} cuadres ({omitidos} ya existían).")
    if nuevos:
        ultimo = max(nuevos, key=lambda c: c.numero_turno)
        print(f"Último turno: {ultimo.numero_turno} — el siguiente cuadre será el {ultimo.numero_turno + 1}.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    asyncio.run(importar(sys.argv[1]))
