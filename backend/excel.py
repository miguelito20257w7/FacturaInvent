"""Exportación a Excel (.xlsx) con openpyxl."""
import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from models import CuadreCaja, Empresa, Producto

_FORMATO_PESOS = '"$"#,##0'


def _autoajustar(ws):
    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ancho + 3, 40)


def cuadres_a_xlsx(cuadres: List[CuadreCaja]) -> io.BytesIO:
    """Mismas columnas que la hoja BASE DE DATOS del Excel original."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BASE DE DATOS"

    encabezados = [
        "TURNO", "FECHA", "HORA", "USUARIO", "JORNADA", "VENTAS NETAS",
        "ENTREGAS", "TARJETAS", "BONOS", "NEQUI O QR", "VENTAS FE.CREDITO",
        "BASE DEL DIA", "BASE ANTERIOR", "TOTAL",
    ]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    for c in cuadres:
        ws.append([
            c.numero_turno,
            c.fecha.isoformat(),
            c.hora.strftime("%I:%M %p"),
            c.usuario_nombre,
            c.jornada,
            c.ventas_netas,
            c.entregas,
            c.tarjetas,
            c.bonos,
            c.nequi_qr,
            c.fact_electronica_credito,
            c.base_del_dia,
            c.base_anterior,
            c.sobrante_faltante,
        ])

    for fila in ws.iter_rows(min_row=2, min_col=6, max_col=14):
        for celda in fila:
            celda.number_format = _FORMATO_PESOS

    _autoajustar(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def inventario_a_xlsx(empresa: Empresa, productos: List[Producto]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.append([f"{empresa.nombre} — NIT {empresa.nit}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    encabezados = [
        "CODIGO FACTURA", "CODIGO BARRAS", "CODIGO INTERNO", "NOMBRE",
        "CANTIDAD", "PRECIO", "PRECIO DIVIDIDO", "PAQUETES", "DESCUENTO",
    ]
    ws.append(encabezados)
    for celda in ws[3]:
        celda.font = Font(bold=True)

    for p in productos:
        ws.append([
            p.codigo_factura,
            p.codigo_barras,
            p.codigo_interno,
            p.nombre,
            p.cantidad_productos,
            p.precio,
            p.precio_dividido,
            p.cantidad_paquetes if p.viene_en_paquetes else "",
            "SI" if p.tiene_descuento else "",
        ])

    for fila in ws.iter_rows(min_row=4, min_col=6, max_col=7):
        for celda in fila:
            celda.number_format = _FORMATO_PESOS

    _autoajustar(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
